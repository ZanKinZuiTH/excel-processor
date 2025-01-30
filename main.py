"""
Excel Document Processor
-----------------------
ระบบประมวลผลและจัดการเอกสาร Excel อัจฉริยะ

ระบบนี้ถูกออกแบบมาเพื่อ:
1. แยกข้อมูลและโครงสร้างของเอกสาร Excel โดยอัตโนมัติ
2. รักษารูปแบบการจัดวางและการจัดรูปแบบต้นฉบับ
3. จัดการระบบเทมเพลตและการพิมพ์

Author: ZanKinZuiTH
Version: 1.0.0
"""

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, Border
from typing import Dict, List, Any, Optional
import json
from sqlalchemy import create_engine, MetaData, Table, Column, String, Integer, JSON
from pathlib import Path
import re
import logging

# ตั้งค่า logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

class ExcelProcessor:
    """
    คลาสหลักสำหรับการประมวลผลเอกสาร Excel
    
    ความสามารถหลัก:
    - แยกข้อมูลและโครงสร้างเอกสาร
    - จัดการรูปแบบและการจัดวาง
    - บันทึกและใช้งานเทมเพลต
    """

    def __init__(self, file_path: str):
        """
        เริ่มต้นระบบประมวลผล Excel
        
        Args:
            file_path: พาธของไฟล์ Excel ที่ต้องการประมวลผล
        """
        self.file_path = file_path
        self.workbook = load_workbook(file_path) if file_path else None
        
        # สร้างการเชื่อมต่อกับฐานข้อมูล
        logger.info(f"กำลังเชื่อมต่อกับฐานข้อมูล...")
        self.engine = create_engine('sqlite:///excel_data.db')
        self.setup_database()

    def setup_database(self):
        """ตั้งค่าโครงสร้างฐานข้อมูล"""
        metadata = MetaData()
        
        # ตารางเก็บข้อมูลเนื้อหา
        self.content_table = Table(
            'Content', metadata,
            Column('entry_id', Integer, primary_key=True),
            Column('sheet_name', String),
            Column('column_name', String),
            Column('cell_value', String)
        )

        # ตารางเก็บข้อมูลโครงสร้าง
        self.structure_table = Table(
            'Structure', metadata,
            Column('entry_id', Integer, primary_key=True),
            Column('sheet_name', String),
            Column('row_number', Integer),
            Column('formatting', JSON)
        )

        # ตารางเก็บเทมเพลต
        self.template_table = Table(
            'Template', metadata,
            Column('template_id', Integer, primary_key=True),
            Column('name', String),
            Column('structure', JSON)
        )

        metadata.create_all(self.engine)
        logger.info("สร้างโครงสร้างฐานข้อมูลเรียบร้อย")

    def extract_customer_info(self, cell_value: str) -> Optional[Dict[str, str]]:
        """
        แยกข้อมูลลูกค้าจากข้อความ
        
        ตัวอย่าง:
        "นางสาว ราตรี สกุลวงษ์" -> {
            "title": "นางสาว",
            "first_name": "ราตรี",
            "last_name": "สกุลวงษ์"
        }
        
        Args:
            cell_value: ข้อความที่ต้องการแยกข้อมูล
            
        Returns:
            Dictionary ของข้อมูลที่แยกแล้ว หรือ None ถ้าไม่สามารถแยกได้
        """
        if not cell_value or not isinstance(cell_value, str):
            return None

        # แยกคำนำหน้า ชื่อ และนามสกุล
        name_match = re.match(r'^(นาย|นาง|นางสาว)\s+(.+?)\s+(.+)$', cell_value)
        if name_match:
            return {
                "title": name_match.group(1),
                "first_name": name_match.group(2),
                "last_name": name_match.group(3)
            }
        return None

    def get_cell_formatting(self, cell) -> dict:
        """
        อ่านข้อมูลการจัดรูปแบบของเซลล์
        
        Args:
            cell: เซลล์ที่ต้องการอ่านข้อมูล
            
        Returns:
            Dictionary ของข้อมูลการจัดรูปแบบ
        """
        font = cell.font
        alignment = cell.alignment
        border = cell.border
        
        return {
            "font": {
                "name": font.name,
                "size": font.size,
                "bold": font.bold,
                "italic": font.italic,
                "color": font.color.rgb if font.color else None
            },
            "alignment": {
                "horizontal": alignment.horizontal,
                "vertical": alignment.vertical,
                "wrap_text": alignment.wrap_text
            },
            "border": {
                "left": border.left.style if border.left else None,
                "right": border.right.style if border.right else None,
                "top": border.top.style if border.top else None,
                "bottom": border.bottom.style if border.bottom else None
            }
        }

    def read_excel_content(self) -> Dict[str, List[Any]]:
        """
        อ่านข้อมูลและการจัดรูปแบบทั้งหมดจากไฟล์ Excel
        
        Returns:
            Dictionary ของข้อมูลและการจัดรูปแบบแยกตาม sheet
        """
        logger.info(f"กำลังอ่านข้อมูลจากไฟล์ {self.file_path}")
        sheet_data = {}
        
        for sheet_name in self.workbook.sheetnames:
            sheet = self.workbook[sheet_name]
            data = []
            formatting = []
            
            for row in sheet.iter_rows():
                row_data = []
                row_formatting = []
                
                for cell in row:
                    cell_value = str(cell.value) if cell.value is not None else ""
                    # ตรวจสอบและแยกข้อมูลลูกค้า
                    customer_info = self.extract_customer_info(cell_value)
                    if customer_info:
                        row_data.append(customer_info)
                    else:
                        row_data.append(cell_value)
                    row_formatting.append(self.get_cell_formatting(cell))
                
                data.append(row_data)
                formatting.append(row_formatting)
            
            sheet_data[sheet_name] = {
                "data": data,
                "formatting": formatting
            }
            
        logger.info(f"อ่านข้อมูลเรียบร้อย พบ {len(sheet_data)} sheets")
        return sheet_data

    def separate_structure_and_content(self, sheet_data: Dict[str, Dict]) -> Dict[str, Dict]:
        """
        แยกข้อมูลและโครงสร้างออกจากกัน
        
        Args:
            sheet_data: ข้อมูลที่อ่านจากไฟล์ Excel
            
        Returns:
            Dictionary ของข้อมูลและโครงสร้างที่แยกแล้ว
        """
        logger.info("กำลังแยกข้อมูลและโครงสร้าง...")
        result = {}
        
        for sheet_name, data in sheet_data.items():
            structure = []
            content = []
            
            headers = data["data"][0] if data["data"] else []
            header_formatting = data["formatting"][0] if data["formatting"] else []
            
            # เก็บข้อมูลส่วนหัว
            structure.append({
                "row_number": 0,
                "type": "header",
                "formatting": header_formatting
            })
            
            for row_idx, (row, row_formatting) in enumerate(zip(data["data"][1:], data["formatting"][1:]), 1):
                row_content = {}
                row_structure = {
                    "row_number": row_idx,
                    "type": "data",
                    "formatting": row_formatting
                }
                
                for header, cell, cell_format in zip(headers, row, row_formatting):
                    if isinstance(cell, dict):  # กรณีเป็นข้อมูลลูกค้าที่แยกแล้ว
                        row_content.update(cell)
                    elif cell and not (isinstance(cell, str) and cell.isspace()):
                        row_content[str(header)] = cell
                
                if row_content:
                    content.append(row_content)
                structure.append(row_structure)
            
            result[sheet_name] = {
                "content": content,
                "structure": structure
            }
        
        logger.info("แยกข้อมูลและโครงสร้างเรียบร้อย")
        return result

    def save_to_database(self, processed_data: Dict[str, Dict]):
        """
        บันทึกข้อมูลลงฐานข้อมูล
        
        Args:
            processed_data: ข้อมูลที่ประมวลผลแล้ว
        """
        logger.info("กำลังบันทึกข้อมูลลงฐานข้อมูล...")
        with self.engine.connect() as conn:
            for sheet_name, data in processed_data.items():
                # บันทึกข้อมูล Content
                for content_row in data["content"]:
                    for col_name, value in content_row.items():
                        if isinstance(value, dict):  # กรณีเป็นข้อมูลลูกค้าที่แยกแล้ว
                            for k, v in value.items():
                                conn.execute(
                                    self.content_table.insert().values(
                                        sheet_name=sheet_name,
                                        column_name=f"{col_name}_{k}",
                                        cell_value=v
                                    )
                                )
                        else:
                            conn.execute(
                                self.content_table.insert().values(
                                    sheet_name=sheet_name,
                                    column_name=col_name,
                                    cell_value=value
                                )
                            )
                
                # บันทึกข้อมูล Structure
                for structure_row in data["structure"]:
                    conn.execute(
                        self.structure_table.insert().values(
                            sheet_name=sheet_name,
                            row_number=structure_row["row_number"],
                            formatting=json.dumps(structure_row["formatting"])
                        )
                    )
        logger.info("บันทึกข้อมูลเรียบร้อย")

    def save_as_template(self, name: str, processed_data: Dict[str, Dict]):
        """
        บันทึกข้อมูลเป็นเทมเพลต
        
        Args:
            name: ชื่อเทมเพลต
            processed_data: ข้อมูลที่จะบันทึกเป็นเทมเพลต
        """
        logger.info(f"กำลังบันทึกเทมเพลต '{name}'...")
        with self.engine.connect() as conn:
            conn.execute(
                self.template_table.insert().values(
                    name=name,
                    structure=json.dumps(processed_data)
                )
            )
        logger.info(f"บันทึกเทมเพลต '{name}' เรียบร้อย")

    def process_file(self) -> Dict[str, Dict]:
        """
        ประมวลผลไฟล์ Excel ทั้งหมด
        
        Returns:
            Dictionary ของข้อมูลที่ประมวลผลแล้ว
        """
        logger.info("เริ่มการประมวลผลไฟล์...")
        sheet_data = self.read_excel_content()
        processed_data = self.separate_structure_and_content(sheet_data)
        self.save_to_database(processed_data)
        logger.info("ประมวลผลไฟล์เสร็จสมบูรณ์")
        return processed_data

def main():
    """ฟังก์ชันหลักสำหรับการทดสอบ"""
    file_path = "นางสาว ราตรี สกุลวงษ์.xlsx"
    if Path(file_path).exists():
        processor = ExcelProcessor(file_path)
        result = processor.process_file()
        
        # บันทึกเป็นเทมเพลต
        processor.save_as_template("customer_template", result)
        
        print(json.dumps(result, ensure_ascii=False, indent=2))
    else:
        logger.error(f"ไม่พบไฟล์ {file_path}")

if __name__ == "__main__":
    main() 