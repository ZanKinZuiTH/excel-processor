"""
Print Management System
---------------------
ระบบจัดการการพิมพ์อัตโนมัติสำหรับเอกสาร Excel

ระบบนี้ถูกออกแบบมาเพื่อ:
1. จัดการคิวการพิมพ์อัตโนมัติ
2. รองรับการพิมพ์แบบกลุ่ม
3. เลือกเครื่องพิมพ์ได้อย่างยืดหยุ่น

Author: ZanKinZuiTH
Version: 1.0.0
"""

import win32print
import win32api
from typing import List, Optional
import json
from pathlib import Path
import tempfile
import os
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
import logging

# ตั้งค่า logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

class PrintManager:
    """
    คลาสหลักสำหรับการจัดการการพิมพ์
    
    ความสามารถหลัก:
    - จัดการคิวการพิมพ์
    - เลือกเครื่องพิมพ์
    - สร้างและจัดรูปแบบเอกสาร
    """

    def __init__(self):
        """เริ่มต้นระบบจัดการการพิมพ์"""
        self.default_printer = win32print.GetDefaultPrinter()
        self.print_queue = []
        logger.info(f"เริ่มต้นระบบพิมพ์ด้วยเครื่องพิมพ์: {self.default_printer}")

    def get_available_printers(self) -> List[str]:
        """
        รับรายชื่อเครื่องพิมพ์ที่มีในระบบ
        
        Returns:
            List ของชื่อเครื่องพิมพ์ที่ใช้งานได้
        """
        printers = []
        for printer in win32print.EnumPrinters(2):
            printers.append(printer[2])
        logger.info(f"พบเครื่องพิมพ์ {len(printers)} เครื่อง")
        return printers

    def set_printer(self, printer_name: str) -> bool:
        """
        เลือกเครื่องพิมพ์ที่จะใช้งาน
        
        Args:
            printer_name: ชื่อเครื่องพิมพ์ที่ต้องการใช้
            
        Returns:
            bool: True ถ้าตั้งค่าสำเร็จ, False ถ้าไม่พบเครื่องพิมพ์
        """
        if printer_name in self.get_available_printers():
            self.default_printer = printer_name
            logger.info(f"เปลี่ยนเครื่องพิมพ์เป็น: {printer_name}")
            return True
        logger.warning(f"ไม่พบเครื่องพิมพ์: {printer_name}")
        return False

    def create_excel_from_data(self, data: dict, template_formatting: Optional[dict] = None) -> str:
        """
        สร้างไฟล์ Excel จากข้อมูลและการจัดรูปแบบ
        
        Args:
            data: ข้อมูลที่จะใส่ในไฟล์
            template_formatting: ข้อมูลการจัดรูปแบบ (ถ้ามี)
            
        Returns:
            str: พาธของไฟล์ Excel ที่สร้าง
        """
        logger.info("กำลังสร้างไฟล์ Excel...")
        wb = Workbook()
        
        for sheet_name, sheet_data in data.items():
            ws = wb.create_sheet(title=sheet_name)
            
            # ใส่ข้อมูลและจัดรูปแบบ
            for row_idx, row_data in enumerate(sheet_data["content"], 1):
                for col_idx, (key, value) in enumerate(row_data.items(), 1):
                    cell = ws.cell(row=row_idx, column=col_idx, value=value)
                    
                    if template_formatting and sheet_name in template_formatting:
                        formatting = template_formatting[sheet_name]["structure"][row_idx-1]["formatting"][col_idx-1]
                        
                        # จัดรูปแบบตัวอักษร
                        font_data = formatting["font"]
                        cell.font = Font(
                            name=font_data["name"],
                            size=font_data["size"],
                            bold=font_data["bold"],
                            italic=font_data["italic"]
                        )
                        
                        # จัดรูปแบบการจัดวาง
                        align_data = formatting["alignment"]
                        cell.alignment = Alignment(
                            horizontal=align_data["horizontal"],
                            vertical=align_data["vertical"],
                            wrap_text=align_data["wrap_text"]
                        )
                        
                        # จัดรูปแบบเส้นขอบ
                        border_data = formatting["border"]
                        cell.border = Border(
                            left=Side(style=border_data["left"]) if border_data["left"] else None,
                            right=Side(style=border_data["right"]) if border_data["right"] else None,
                            top=Side(style=border_data["top"]) if border_data["top"] else None,
                            bottom=Side(style=border_data["bottom"]) if border_data["bottom"] else None
                        )
        
        if "Sheet" in wb.sheetnames:
            wb.remove(wb["Sheet"])
        
        # บันทึกไฟล์ชั่วคราว
        temp_file = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
        wb.save(temp_file.name)
        logger.info(f"สร้างไฟล์ Excel เรียบร้อย: {temp_file.name}")
        return temp_file.name

    def add_to_queue(self, data: dict, template_formatting: Optional[dict] = None):
        """
        เพิ่มงานพิมพ์เข้าคิว
        
        Args:
            data: ข้อมูลที่จะพิมพ์
            template_formatting: ข้อมูลการจัดรูปแบบ (ถ้ามี)
        """
        excel_file = self.create_excel_from_data(data, template_formatting)
        self.print_queue.append(excel_file)
        logger.info(f"เพิ่มงานพิมพ์เข้าคิว: {excel_file}")

    def print_all(self):
        """พิมพ์งานทั้งหมดในคิว"""
        logger.info(f"เริ่มพิมพ์งานในคิว ({len(self.print_queue)} งาน)")
        for file_path in self.print_queue:
            try:
                win32api.ShellExecute(0, "print", file_path, f'/d:"{self.default_printer}"', ".", 0)
                logger.info(f"พิมพ์ไฟล์: {file_path}")
            except Exception as e:
                logger.error(f"เกิดข้อผิดพลาดในการพิมพ์: {str(e)}")
            finally:
                # ลบไฟล์ชั่วคราวหลังจากส่งไปพิมพ์
                os.unlink(file_path)
                logger.debug(f"ลบไฟล์ชั่วคราว: {file_path}")
        self.print_queue.clear()
        logger.info("พิมพ์งานทั้งหมดเสร็จสิ้น")

    def print_file(self, file_path: str) -> bool:
        """
        พิมพ์ไฟล์ทันที
        
        Args:
            file_path: พาธของไฟล์ที่ต้องการพิมพ์
            
        Returns:
            bool: True ถ้าพิมพ์สำเร็จ, False ถ้าไม่พบไฟล์
        """
        if Path(file_path).exists():
            try:
                win32api.ShellExecute(0, "print", file_path, f'/d:"{self.default_printer}"', ".", 0)
                logger.info(f"พิมพ์ไฟล์: {file_path}")
                return True
            except Exception as e:
                logger.error(f"เกิดข้อผิดพลาดในการพิมพ์: {str(e)}")
                return False
        logger.warning(f"ไม่พบไฟล์: {file_path}")
        return False 